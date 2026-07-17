#!/usr/bin/env python3
"""Build master FrankX blog article registry (CSV/JSON/MD/XLSX)."""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent
PUBLIC = ROOT / "public"
BLOG = ROOT / "content" / "blog"


def file_exists(p: str | None) -> bool:
    if not p:
        return False
    return (PUBLIC / str(p).lstrip("/")).exists()


def main() -> None:
    articles = json.loads((OUT / "article-registry-2026-07-15.json").read_text(encoding="utf-8"))
    git_rows = json.loads((OUT / "git-attribution-2026-07-15.json").read_text(encoding="utf-8"))
    git = {r["slug"]: r for r in git_rows}

    vs_data = json.loads((ROOT / "data" / "blog-visual-system.json").read_text(encoding="utf-8"))
    vs = {p["slug"]: p for p in vs_data.get("posts", [])}

    for a in articles:
        slug = a["slug"].split("/")[-1]
        g = git.get(slug, {})
        a["last_git_author"] = g.get("last_author", "")
        a["last_git_date"] = g.get("last_date", "")
        a["last_git_subject"] = g.get("last_subject", "")
        a["agent_hint"] = g.get("agent_hint", "unknown")
        a["recent_commits"] = g.get("recent_commits", 0)

        # refresh vs if missing
        if not a.get("vs_hero") and slug in vs:
            v = vs[slug]
            a["vs_hero"] = v.get("hero", "")
            a["vs_prompt"] = v.get("prompt", "")
            a["vs_reviewStatus"] = v.get("reviewStatus", "")
            a["vs_generationMode"] = v.get("generationMode", "")
            a["vs_infographic"] = v.get("infographic", "")

        hc = a.get("header_class", "")
        img_ok = bool(a.get("resolved_hero")) and file_exists(a.get("resolved_hero"))

        if a.get("status_bucket") == "draft_folder":
            pub = "draft"
        else:
            pub = "published_tree"

        content_review = "not_recorded"
        visual_review = a.get("vs_reviewStatus") or (
            "pass_heuristic" if hc in ("premium_raster", "raster_other") and img_ok else "needs_review"
        )
        if hc == "svg_or_visual_system":
            visual_review = "fail_svg_slop_risk"
        if hc == "missing":
            visual_review = "fail_missing_hero"

        seo_score = 0
        if a.get("title"):
            seo_score += 1
        if a.get("description") and len(str(a.get("description"))) > 40:
            seo_score += 1
        if a.get("keywords"):
            seo_score += 1
        if a.get("category"):
            seo_score += 1
        if int(a.get("internal_link_count") or 0) >= 3:
            seo_score += 1
        if img_ok:
            seo_score += 1

        prompt = a.get("vs_prompt") or ""
        featured_flag = a.get("featured") in (True, "true", "True", "yes", "1")

        if hc == "missing":
            action = "generate_hero+wire_frontmatter"
            prio = 30
        elif hc == "svg_or_visual_system":
            action = "replace_svg_header"
            prio = 25
        else:
            action = "audit_content_accuracy"
            prio = 0

        prio += 10 if not prompt else 0
        prio += 10 if seo_score < 4 else 0
        prio += 5 if int(a.get("word_count") or 0) > 2500 else 0
        prio += 5 if featured_flag else 0
        if a.get("status_bucket") == "draft_folder":
            action = "draft_complete_or_archive"
            prio = max(prio - 20, 0)

        a.update(
            {
                "publish_status": pub,
                "content_review_status": content_review,
                "visual_review_status": visual_review,
                "prompt_used": prompt,
                "hero_path_final": a.get("resolved_hero") or a.get("image") or "",
                "images_in_article": a.get("body_image_count", 0),
                "seo_completeness_6": seo_score,
                "priority_score": prio,
                "action_needed": action,
                "registry_updated_at": datetime.now(timezone.utc).isoformat(),
            }
        )

    articles.sort(key=lambda x: (-int(x.get("priority_score") or 0), x.get("slug", "")))
    (OUT / "article-registry-master.json").write_text(
        json.dumps(articles, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    fields = [
        "slug",
        "title",
        "date",
        "publish_status",
        "author",
        "category",
        "featured",
        "agent_hint",
        "last_git_author",
        "last_git_date",
        "last_git_subject",
        "hero_path_final",
        "header_class",
        "image_file_exists",
        "vs_hero",
        "vs_infographic",
        "prompt_used",
        "visual_review_status",
        "content_review_status",
        "word_count",
        "images_in_article",
        "body_svg_count",
        "internal_link_count",
        "seo_completeness_6",
        "priority_score",
        "action_needed",
        "description",
        "file",
    ]

    with (OUT / "article-registry-master.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for a in articles:
            row = {k: a.get(k, "") for k in fields}
            if isinstance(row.get("prompt_used"), str):
                row["prompt_used"] = row["prompt_used"].replace("\n", " ").strip()
            w.writerow(row)

    summary = [
        ("generated_at", datetime.now(timezone.utc).isoformat()),
        ("total_articles", len(articles)),
        ("published_tree", sum(1 for a in articles if a.get("publish_status") == "published_tree")),
        ("drafts", sum(1 for a in articles if a.get("publish_status") == "draft")),
        ("missing_hero", sum(1 for a in articles if a.get("header_class") == "missing")),
        ("svg_slop_risk", sum(1 for a in articles if a.get("header_class") == "svg_or_visual_system")),
        ("premium_raster", sum(1 for a in articles if a.get("header_class") == "premium_raster")),
        ("raster_other", sum(1 for a in articles if a.get("header_class") == "raster_other")),
        ("visual_system_tracked", sum(1 for a in articles if a.get("vs_hero"))),
        (
            "avg_words",
            int(sum(int(a.get("word_count") or 0) for a in articles) / max(1, len(articles))),
        ),
        (
            "action_generate_hero",
            sum(1 for a in articles if a.get("action_needed") == "generate_hero+wire_frontmatter"),
        ),
        (
            "action_replace_svg",
            sum(1 for a in articles if a.get("action_needed") == "replace_svg_header"),
        ),
        (
            "action_audit_content",
            sum(1 for a in articles if a.get("action_needed") == "audit_content_accuracy"),
        ),
    ]

    # XLSX
    xlsx_path = OUT / "article-registry-master.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registry"
        ws.append(fields)
        header_fill = PatternFill("solid", fgColor="0B1F1A")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        fills = {
            "generate_hero+wire_frontmatter": PatternFill("solid", fgColor="FECACA"),
            "replace_svg_header": PatternFill("solid", fgColor="FDE68A"),
            "audit_content_accuracy": PatternFill("solid", fgColor="BFDBFE"),
            "draft_complete_or_archive": PatternFill("solid", fgColor="E5E7EB"),
            "maintain": PatternFill("solid", fgColor="D1FAE5"),
        }
        action_col = fields.index("action_needed") + 1
        for a in articles:
            row = []
            for k in fields:
                v = a.get(k, "")
                if isinstance(v, str):
                    v = v.replace("\n", " ")
                row.append(v)
            ws.append(row)
            fill = fills.get(a.get("action_needed"))
            if fill:
                ws.cell(row=ws.max_row, column=action_col).fill = fill

        ws2 = wb.create_sheet("Summary")
        ws2.append(["metric", "value"])
        for k, v in summary:
            ws2.append([k, v])

        ws3 = wb.create_sheet("PriorityHeaders")
        pri_fields = [
            "slug",
            "title",
            "header_class",
            "action_needed",
            "priority_score",
            "hero_path_final",
            "prompt_used",
        ]
        ws3.append(pri_fields)
        for a in articles:
            if a.get("action_needed") in (
                "generate_hero+wire_frontmatter",
                "replace_svg_header",
            ):
                ws3.append(
                    [
                        (
                            a.get(k, "").replace("\n", " ")
                            if isinstance(a.get(k, ""), str)
                            else a.get(k, "")
                        )
                        for k in pri_fields
                    ]
                )
        wb.save(xlsx_path)
        print("XLSX", xlsx_path)
    except Exception as e:
        print("xlsx failed", e)

    # Markdown
    lines: list[str] = []
    lines.append("# FrankX Blog Article Registry\n\n")
    lines.append(
        f"**Generated:** {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  \n"
    )
    lines.append("**Source repo:** `frankxai/frankx.ai-vercel-website` branch `codex/blog`  \n")
    lines.append("**Content path:** `content/blog/*.mdx`  \n")
    lines.append(f"**Total articles inventoried:** {len(articles)}  \n\n")
    lines.append("## Executive snapshot\n\n")
    lines.append("| Metric | Count |\n|---|---:|\n")
    for k, v in summary:
        lines.append(f"| {k} | {v} |\n")
    lines.append("\n## Where everything lives\n\n")
    lines.append("| Layer | Path |\n|---|---|\n")
    lines.append("| Production site + MDX SSOT | `starlight/repos/frankx.ai-vercel-website` |\n")
    lines.append("| Private authoring twin | `starlight/repos/FrankX` (does not deploy) |\n")
    lines.append("| Empire content strategy | `starlight/CONTENT-STRATEGY.md` |\n")
    lines.append("| Repo content strategy | `CONTENT_STRATEGY.md`, `docs/CONTENT_ROADMAP.md` |\n")
    lines.append("| Visual OS manifests | `data/blog-visual-system.json`, `data/blog-heroes.json` |\n")
    lines.append("| Blog images | `public/images/blog/` (+ `generated/`, `visual-system/`) |\n")
    lines.append("| This registry | `data/blog-ops/article-registry-master.*` |\n\n")
    lines.append("## Priority header queue (SVG slop + missing)\n\n")
    lines.append("| Priority | Slug | Class | Action | Current hero |\n|---:|---|---|---|---|\n")
    for a in articles:
        if a.get("action_needed") in (
            "generate_hero+wire_frontmatter",
            "replace_svg_header",
        ):
            lines.append(
                f"| {a.get('priority_score')} | `{a.get('slug')}` | {a.get('header_class')} | {a.get('action_needed')} | `{a.get('hero_path_final', '')}` |\n"
            )
    lines.append("\n## Full registry (compact)\n\n")
    lines.append(
        "| Slug | Date | Agent hint | Header | Visual review | SEO/6 | Words | Action |\n|---|---|---|---|---|---:|---:|---|\n"
    )
    for a in articles:
        if a.get("publish_status") == "draft":
            continue
        lines.append(
            f"| `{a.get('slug')}` | {str(a.get('date', ''))[:10]} | {a.get('agent_hint')} | {a.get('header_class')} | {a.get('visual_review_status')} | {a.get('seo_completeness_6')} | {a.get('word_count')} | {a.get('action_needed')} |\n"
        )
    lines.append("\n## Field definitions\n\n")
    lines.append(
        "- **agent_hint**: derived from last 5 git commits (subject/author heuristics). Most commits are authored as `Frank`; true agent lineage is often not in git metadata — fill `agent_author_meta` in frontmatter going forward.\n"
    )
    lines.append(
        "- **content_review_status**: `not_recorded` until a second-agent accuracy pass logs evidence.\n"
    )
    lines.append(
        "- **visual_review_status**: from visual-system manifest or header class heuristics.\n"
    )
    lines.append("- **prompt_used**: from `data/blog-visual-system.json` when present.\n")
    lines.append(
        "- **action_needed**: operational next step for multi-agent upgrade swarm.\n\n"
    )
    lines.append("## Operating rules for scale\n\n")
    lines.append(
        "1. Every new/updated post MUST log: authoring agent, reviewer agent, hero path, prompt path, visual QA score, content accuracy pass, SEO/AEO checklist.\n"
    )
    lines.append(
        "2. Headers: Tier B cinematic raster 16:9 preferred; ban decorative SVG as primary hero.\n"
    )
    lines.append(
        "3. Diagrams/infographics: exact-text code/SVG only when labels matter; never as OG/hero.\n"
    )
    lines.append("4. Maker ≠ checker for accuracy claims and tool docs currency.\n")

    md_path = OUT / "ARTICLE-REGISTRY.md"
    md_path.write_text("".join(lines), encoding="utf-8")
    print("MD", md_path)
    print("CSV", OUT / "article-registry-master.csv")
    print(
        "priority",
        sum(
            1
            for a in articles
            if a.get("action_needed")
            in ("generate_hero+wire_frontmatter", "replace_svg_header")
        ),
    )
    print("top12:")
    for a in articles[:12]:
        print(a["priority_score"], a["slug"], a["action_needed"], a["header_class"])


if __name__ == "__main__":
    main()
